#!/usr/bin/env python3
"""
Parse PRE MPLS circuit Excel spreadsheets into a plain-text report.

Supports both layouts:
  - Full:  Full/Partial, Comments, Circuit #, New Circuit #, Circuit Name, email, sites..., AC, MPLS
  - Short: Circuit #, New Circuit #, Circuit Name, sites..., AC, MPLS

Site columns are detected by position (after Circuit Name, before AC/MPLS).

Usage:
    pip install openpyxl
    python parse_mpls_circuits.py input.xlsx output.txt
    python parse_mpls_circuits.py input.xlsx --diagnose
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.worksheet.worksheet import Worksheet

PORT_RE = re.compile(r"\d+/\d+/\d+")

# Default Excel theme accent palette (Office 2007+)
THEME_PALETTE = [
    "FFFFFF", "000000", "E7E6E6", "44546A", "4472C4", "ED7D31",
    "A5A5A5", "FFC000", "5B9BD5", "70AD47", "0563C1", "954F72",
]


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class SitePort:
    site: str
    port: str
    color: str | None


@dataclass
class Connection:
    endpoints: list[SitePort]


@dataclass
class Circuit:
    number: str
    new_number: str
    name: str
    connections: list[Connection] = field(default_factory=list)


@dataclass
class SheetLayout:
    header_row: int
    site_header_row: int
    first_data_row: int
    circuit: int | None
    new_circuit: int | None
    name: int | None
    email: int | None
    ac: int | None
    mpls: int | None
    sites: list[tuple[int, str]]


# ---------------------------------------------------------------------------
# Color
# ---------------------------------------------------------------------------

KNOWN_COLORS = {
    "FF90C060": "light green",
    "FF92D050": "light green",
    "FF2E7D32": "dark green",
    "FF00695C": "teal",
    "FF1565C0": "blue",
    "FF4472C4": "blue",
    "FF5B9BD5": "blue",
    "FF8DB4E2": "light blue",
    "FFB4C6E7": "light blue",
    "FF8EAADB": "light blue",
    "FF00AA00": "green",
    "FF008000": "green",
    "FF006400": "dark green",
    "FF70AD47": "green",
    "FF548235": "dark green",
    "FF375623": "dark green",
    "FF7030A0": "purple",
    "FF000000": "black",
    "FF0000FF": "blue",
    "FF00FF00": "green",
    "FFC6EFCE": "light green",
    "FFE2EFDA": "light green",
}


def _hex_to_rgb(hex6: str) -> tuple[int, int, int]:
    h = hex6[-6:]
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _rgb_to_hex6(r: int, g: int, b: int) -> str:
    return f"{r:02X}{g:02X}{b:02X}"


def _apply_tint(hex6: str, tint: float) -> str:
    r, g, b = _hex_to_rgb(hex6)
    if tint < 0:
        factor = 1 + tint
        r, g, b = int(r * factor), int(g * factor), int(b * factor)
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    return _rgb_to_hex6(r, g, b)


def _normalize_hex(raw: str) -> str | None:
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip().upper()
    if raw in ("00000000", "FF000000", "FFFFFFFF", "00FFFFFF", "AUTO", "NONE"):
        return None
    if len(raw) == 8:
        return raw
    if len(raw) == 6:
        return "FF" + raw
    return None


def _classify_hex6(hex6: str) -> str:
    """Guess a color name from RGB when not in KNOWN_COLORS."""
    r, g, b = _hex_to_rgb(hex6)
    if b > max(r, g) + 20 and b > 120:
        return "light blue" if b > 150 or (b > 100 and g > 100) else "blue"
    if g > max(r, b) + 20 and g > 80:
        if g < 110 or (r < 80 and b < 80):
            return "dark green"
        if g > 170:
            return "light green"
        return "green"
    return f"#{hex6[-6:]}"


def _name_from_hex(hex8: str) -> str:
    key = hex8 if len(hex8) == 8 else "FF" + hex8
    if key in KNOWN_COLORS:
        return KNOWN_COLORS[key]
    hex6 = key[-6:]
    if hex6 in KNOWN_COLORS:
        return KNOWN_COLORS[hex6]
    return _classify_hex6(hex6)
    

def resolve_color(color) -> str | None:
    """Resolve an openpyxl Color object to a human-readable color name."""
    if color is None:
        return None
    ctype = getattr(color, "type", None)

    if ctype == "rgb":
        try:
            raw = color.rgb
        except Exception:
            raw = None
        hex8 = _normalize_hex(raw) if raw else None
        return _name_from_hex(hex8) if hex8 else None

    if ctype == "indexed":
        idx = color.indexed
        if idx is not None and 0 <= idx < len(COLOR_INDEX):
            hex8 = _normalize_hex(COLOR_INDEX[idx])
            return _name_from_hex(hex8) if hex8 else None

    if ctype == "theme":
        theme = color.theme
        tint = color.tint or 0.0
        if theme is not None and 0 <= theme < len(THEME_PALETTE):
            hex6 = _apply_tint(THEME_PALETTE[theme], tint)
            return _name_from_hex("FF" + hex6)

    return None


def cell_fill_color(cell) -> str | None:
    try:
        fill = cell.fill
        if not fill or fill.fill_type in (None, "none"):
            return None
        for attr in ("fgColor", "start_color", "bgColor"):
            c = getattr(fill, attr, None)
            name = resolve_color(c)
            if name:
                return name
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Merge index
# ---------------------------------------------------------------------------

class MergeIndex:
    def __init__(self, sheet: Worksheet):
        self._top_left: dict[tuple[int, int], tuple[int, int]] = {}
        self._range: dict[tuple[int, int], tuple[int, int, int, int]] = {}
        for merged in sheet.merged_cells.ranges:
            tl = (merged.min_row, merged.min_col)
            bounds = (merged.min_row, merged.max_row, merged.min_col, merged.max_col)
            self._range[tl] = bounds
            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    self._top_left[(r, c)] = tl

    def top_left(self, row: int, col: int) -> tuple[int, int] | None:
        return self._top_left.get((row, col))

    def bounds(self, row: int, col: int) -> tuple[int, int, int, int] | None:
        tl = self.top_left(row, col)
        return self._range.get(tl) if tl else None


def merged_value(sheet: Worksheet, merges: MergeIndex, row: int, col: int):
    tl = merges.top_left(row, col)
    if tl:
        return sheet.cell(*tl).value
    return sheet.cell(row, col).value


# ---------------------------------------------------------------------------
# Column / header detection
# ---------------------------------------------------------------------------

def classify_header(text: str) -> str | None:
    lower = text.lower().strip()
    if not lower:
        return None
    if re.fullmatch(r"full\s*/\s*partial\s*\??", lower):
        return "full_partial"
    if re.fullmatch(r"comments\s*:?", lower):
        return "comments"
    if re.fullmatch(r"new\s+circuit\s*#?\s*:?", lower):
        return "new_circuit"
    if re.fullmatch(r"circuit\s*#\s*:?", lower):
        return "circuit"
    if re.fullmatch(r"circuit\s+name\s*:?", lower):
        return "name"
    if re.search(r"email|noc|cau", lower) and len(lower) < 40:
        return "email"
    if re.fullmatch(r"ac\s*:?", lower):
        return "ac"
    if ("mpls" in lower or "mux" in lower or "port reuse" in lower) and len(lower) < 60:
        return "mpls"
    return None


def find_header_row(sheet: Worksheet) -> int:
    for row in sheet.iter_rows(max_row=min(10, sheet.max_row)):
        for cell in row:
            kind = classify_header(str(cell.value or ""))
            if kind in ("full_partial", "circuit", "new_circuit", "name"):
                return cell.row
    return 1


def header_text(sheet: Worksheet, row: int, col: int) -> str:
    return str(sheet.cell(row, col).value or "").strip()


def detect_layout(sheet: Worksheet) -> SheetLayout:
    header_row = find_header_row(sheet)

    fixed: dict[str, int] = {}
    for cell in sheet[header_row]:
        kind = classify_header(str(cell.value or ""))
        if kind and kind not in fixed:
            fixed[kind] = cell.column

    site_header_row = header_row
    next_row = header_row + 1
    if next_row <= sheet.max_row:
        meta_end = max(
            fixed.get(k, 0) for k in ("circuit", "new_circuit", "name", "email")
        )
        count_next = sum(
            1
            for c in range(meta_end + 1, sheet.max_column + 1)
            if header_text(sheet, next_row, c)
            and not classify_header(header_text(sheet, next_row, c))
        )
        count_main = sum(
            1
            for c in range(meta_end + 1, sheet.max_column + 1)
            if header_text(sheet, header_row, c)
            and not classify_header(header_text(sheet, header_row, c))
        )
        if count_next > count_main:
            site_header_row = next_row

    stop = min(
        (fixed[c] for c in ("ac", "mpls") if c in fixed),
        default=sheet.max_column + 1,
    )
    meta_end = max(
        fixed.get(k, 0) for k in ("circuit", "new_circuit", "name", "email")
    )
    fixed_cols = set(fixed.values())

    sites: list[tuple[int, str]] = []
    for col in range(meta_end + 1, stop):
        if col in fixed_cols:
            continue
        name = header_text(sheet, site_header_row, col)
        if not name:
            name = header_text(sheet, header_row, col)
        if name and not classify_header(name):
            sites.append((col, name))

    if not sites:
        sites = infer_site_columns_from_data(sheet, meta_end + 1, header_row + 2)

    if not sites:
        raise ValueError(
            "No site columns found. Expected location/site columns after "
            "Circuit Name and before AC/MPLS. Run with --diagnose."
        )

    first_data_row = header_row + 1
    if site_header_row > header_row:
        first_data_row = site_header_row + 1

    while first_data_row <= sheet.max_row:
        circ_col = fixed.get("circuit")
        name_col = fixed.get("name")
        circ = header_text(sheet, first_data_row, circ_col) if circ_col else ""
        name = header_text(sheet, first_data_row, name_col) if name_col else ""
        if circ or name:
            if classify_header(circ) or classify_header(name):
                first_data_row += 1
                continue
            break
        first_data_row += 1

    return SheetLayout(
        header_row=header_row,
        site_header_row=site_header_row,
        first_data_row=first_data_row,
        circuit=fixed.get("circuit"),
        new_circuit=fixed.get("new_circuit"),
        name=fixed.get("name"),
        email=fixed.get("email"),
        ac=fixed.get("ac"),
        mpls=fixed.get("mpls"),
        sites=sites,
    )


def infer_site_columns_from_data(
    sheet: Worksheet, start_col: int, start_row: int
) -> list[tuple[int, str]]:
    found: list[tuple[int, str]] = []
    for col in range(start_col, sheet.max_column + 1):
        header = header_text(sheet, start_row - 1, col) or header_text(
            sheet, start_row - 2, col
        )
        if header and classify_header(header):
            continue
        for row in range(start_row, min(start_row + 200, sheet.max_row + 1)):
            val = sheet.cell(row, col).value
            if val and PORT_RE.search(str(val)):
                label = header or f"Col{col}"
                found.append((col, label))
                break
    return found


# ---------------------------------------------------------------------------
# Port extraction
# ---------------------------------------------------------------------------

def _cell_raw_value(
    values_sheet: Worksheet, styles_sheet: Worksheet, row: int, col: int
) -> str | None:
    for sheet in (values_sheet, styles_sheet):
        val = sheet.cell(row, col).value
        if val is not None:
            text = str(val).strip()
            if text and not text.startswith("="):
                return text
    return None


def port_at_row(
    values_sheet: Worksheet,
    styles_sheet: Worksheet,
    merges: MergeIndex,
    row: int,
    col: int,
    block_start: int = 0,
) -> tuple[str | None, str | None]:
    """Return port text and fill color for one site cell.

    Handles: direct values, vertically merged multiline cells, and multiline
    text stored in an earlier row while later rows carry only the fill color.
    """
    style_cell = styles_sheet.cell(row, col)
    color = cell_fill_color(style_cell)

    bounds = merges.bounds(row, col)
    if bounds and bounds[1] > bounds[0]:
        top_row = bounds[0]
        raw = _cell_raw_value(values_sheet, styles_sheet, top_row, col)
        if raw:
            lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
            idx = row - top_row
            if idx < len(lines):
                if color is None:
                    color = cell_fill_color(styles_sheet.cell(top_row, col))
                return lines[idx], color
        return None, None

    # Multiline text in an earlier row of the same column (common MPLS layout).
    if block_start:
        for src_row in range(block_start, row + 1):
            raw = _cell_raw_value(values_sheet, styles_sheet, src_row, col)
            if raw and "\n" in raw:
                lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
                idx = row - src_row
                if idx < len(lines):
                    return lines[idx], color

    raw = _cell_raw_value(values_sheet, styles_sheet, row, col)
    if raw:
        if "\n" in raw:
            lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
            idx = row - block_start if block_start else 0
            if idx < len(lines):
                return lines[idx], color
            return lines[0], color
        return raw, color

    return None, None


# ---------------------------------------------------------------------------
# Circuit parsing
# ---------------------------------------------------------------------------

def read_field(
    sheet: Worksheet,
    merges: MergeIndex,
    row: int,
    col: int | None,
    formula_sheet: Worksheet | None = None,
    circuit_col: int | None = None,
) -> str:
    if col is None:
        return ""
    val = merged_value(sheet, merges, row, col)
    if (val is None or (isinstance(val, str) and val.startswith("="))) and formula_sheet:
        raw = merged_value(formula_sheet, merges, row, col)
        if isinstance(raw, str) and raw.startswith("="):
            val = _eval_formula(raw, sheet, merges, circuit_col or 3)
        elif raw is not None:
            val = raw
    elif isinstance(val, str) and val.startswith("="):
        val = _eval_formula(val, sheet, merges, circuit_col or 3)
    return str(val or "").strip()


NEW_CIRCUIT_FORMULA = re.compile(
    r'=LEFT\(([A-Z]+)(\d+),LEN\(\1\2\)-2\)\s*&\s*"3"\s*&\s*RIGHT\(\1\2,1\)',
    re.I,
)


def _col_letters_to_num(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _eval_formula(formula: str, sheet: Worksheet, merges: MergeIndex, circuit_col: int) -> str:
    m = NEW_CIRCUIT_FORMULA.match(formula.strip())
    if m:
        ref_col = _col_letters_to_num(m.group(1))
        ref_row = int(m.group(2))
        base = str(merged_value(sheet, merges, ref_row, ref_col) or "")
        if len(base) >= 2:
            return base[:-2] + "3" + base[-1]
    return formula


def row_has_site_ports(
    values_sheet: Worksheet,
    styles_sheet: Worksheet,
    merges: MergeIndex,
    row: int,
    sites: list[tuple[int, str]],
    block_start: int = 0,
) -> bool:
    for col, _ in sites:
        port, _ = port_at_row(
            values_sheet, styles_sheet, merges, row, col, block_start
        )
        if port:
            return True
    return False


def block_end(
    values_sheet: Worksheet,
    styles_sheet: Worksheet,
    merges: MergeIndex,
    start: int,
    circuit_col: int | None,
    sites: list[tuple[int, str]] | None = None,
) -> int:
    end = start
    if circuit_col:
        for row in range(start + 1, values_sheet.max_row + 1):
            tl = merges.top_left(row, circuit_col)
            if tl and tl[0] == start:
                end = row
            else:
                break
    if sites:
        for row in range(end + 1, values_sheet.max_row + 1):
            circ = str(values_sheet.cell(row, circuit_col).value or "").strip() if circuit_col else ""
            if circ:
                break
            if row_has_site_ports(
                values_sheet, styles_sheet, merges, row, sites, start
            ):
                end = row
            else:
                break
    return end


def is_circuit_start(
    sheet: Worksheet, merges: MergeIndex, row: int, layout: SheetLayout
) -> bool:
    if layout.circuit:
        tl = merges.top_left(row, layout.circuit)
        if tl and tl[0] != row:
            return False
    circ = read_field(sheet, merges, row, layout.circuit)
    name = read_field(sheet, merges, row, layout.name)
    if not circ and not name:
        return False
    if classify_header(circ) or classify_header(name):
        return False
    return True


def connections_in_block(
    values_sheet: Worksheet,
    styles_sheet: Worksheet,
    merges: MergeIndex,
    start: int,
    end: int,
    sites: list[tuple[int, str]],
) -> list[Connection]:
    result: list[Connection] = []

    for row in range(start, end + 1):
        ports: list[SitePort] = []
        for col, site_name in sites:
            port, color = port_at_row(
                values_sheet, styles_sheet, merges, row, col, start
            )
            if port:
                ports.append(SitePort(site_name, port, color))

        if len(ports) >= 2:
            for i in range(0, len(ports) - 1, 2):
                if i + 1 < len(ports):
                    result.append(Connection([ports[i], ports[i + 1]]))
        elif len(ports) == 1:
            result.append(Connection([ports[0]]))

    return result


def parse_circuits(
    values_sheet: Worksheet,
    styles_sheet: Worksheet,
    formula_sheet: Worksheet | None = None,
) -> tuple[list[Circuit], SheetLayout]:
    layout = detect_layout(values_sheet)
    merges = MergeIndex(values_sheet)
    formula_sheet = formula_sheet or values_sheet
    circuits: list[Circuit] = []
    row = layout.first_data_row

    while row <= values_sheet.max_row:
        if not is_circuit_start(values_sheet, merges, row, layout):
            row += 1
            continue

        end = block_end(
            values_sheet, styles_sheet, merges, row, layout.circuit, layout.sites
        )
        conns = connections_in_block(
            values_sheet, styles_sheet, merges, row, end, layout.sites
        )

        circuits.append(
            Circuit(
                number=read_field(
                    values_sheet, merges, row, layout.circuit, formula_sheet, layout.circuit
                ),
                new_number=read_field(
                    values_sheet, merges, row, layout.new_circuit, formula_sheet, layout.circuit
                ),
                name=read_field(
                    values_sheet, merges, row, layout.name, formula_sheet, layout.circuit
                ),
                connections=conns,
            )
        )
        row = end + 1

    return circuits, layout


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------

def fmt_site(sp: SitePort) -> str:
    color = f", color: {sp.color}" if sp.color else ""
    return f"{sp.site} (port: {sp.port}{color})"


def render(circuits: list[Circuit]) -> str:
    blocks: list[str] = []

    for n, circ in enumerate(circuits, 1):
        lines = [
            f"Block {n}",
            f"1. Circuit #:     {circ.number or '(blank)'}",
            f"2. New Circuit #: {circ.new_number or '(blank)'}",
            f"3. Circuit Name:  {circ.name or '(blank)'}",
        ]
        if not circ.connections:
            lines.append("   (No site connections found)")
        else:
            for idx, conn in enumerate(circ.connections, 1):
                lines.append(f"   Connection {idx}:")
                for sp in conn.endpoints:
                    lines.append(f"      {fmt_site(sp)}")
                if idx < len(circ.connections):
                    lines.append("")
        blocks.append("\n".join(lines))

    return "\n\n".join(blocks) + "\n"


# ---------------------------------------------------------------------------
# Diagnose
# ---------------------------------------------------------------------------

def load_sheets(path: Path) -> tuple[Worksheet, Worksheet, Worksheet]:
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_styles = openpyxl.load_workbook(path, data_only=False)
    values = pick_sheet(wb_values)
    styles = pick_sheet(wb_styles)
    return values, styles, styles


def diagnose(values: Worksheet, styles: Worksheet) -> None:
    print(f"Sheet : {values.title!r}")
    print(f"Size  : {values.max_row} rows x {values.max_column} cols")

    try:
        layout = detect_layout(values)
    except ValueError as exc:
        print(f"\nLayout error: {exc}")
        print("\nFirst 5 rows (non-empty cells):")
        for r in range(1, min(6, values.max_row + 1)):
            cells = [(c.column, repr(c.value)) for c in values[r] if c.value is not None]
            if cells:
                print(f"  row {r}: {cells[:20]}{'...' if len(cells) > 20 else ''}")
        return

    print(f"\nHeader row       : {layout.header_row}")
    print(f"Site header row  : {layout.site_header_row}")
    print(f"First data row   : {layout.first_data_row}")
    print(f"Circuit col      : {layout.circuit}")
    print(f"New circuit col  : {layout.new_circuit}")
    print(f"Name col         : {layout.name}")
    print(f"Email col        : {layout.email}")
    print(f"AC col           : {layout.ac}")
    print(f"MPLS col         : {layout.mpls}")
    print(f"Site columns     : {len(layout.sites)}")
    for col, name in layout.sites[:10]:
        print(f"    col {col}: {name!r}")
    if len(layout.sites) > 10:
        print(f"    ... and {len(layout.sites) - 10} more")

    circuits, _ = parse_circuits(values, styles, styles)
    print(f"\nCircuits found   : {len(circuits)}")
    if circuits:
        c = circuits[0]
        print(f"\nFirst circuit:")
        print(f"  number : {c.number!r}")
        print(f"  name   : {c.name!r}")
        print(f"  conns  : {len(c.connections)}")
        for conn in c.connections[:3]:
            for sp in conn.endpoints:
                print(f"    - {sp.site}: {sp.port} ({sp.color})")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def pick_sheet(wb: openpyxl.Workbook) -> Worksheet:
    for name in wb.sheetnames:
        if "mpls" in name.lower() or "pre" in name.lower():
            return wb[name]
    return wb.active


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Parse PRE MPLS circuit Excel file.")
    parser.add_argument("input", help="Path to .xlsx file")
    parser.add_argument("output", nargs="?", help="Output .txt path (default: <input>_output.txt)")
    parser.add_argument("--diagnose", action="store_true", help="Print sheet structure and exit")
    args = parser.parse_args(argv)

    path = Path(args.input)
    if not path.exists():
        print(f"Error: file not found: {path}", file=sys.stderr)
        return 1

    values, styles, formulas = load_sheets(path)

    if args.diagnose:
        diagnose(values, styles)
        return 0

    circuits, layout = parse_circuits(values, styles, formulas)
    text = render(circuits)

    out_path = Path(args.output) if args.output else path.with_name(path.stem + "_output.txt")
    out_path.write_text(text, encoding="utf-8")

    print(f"Parsed {len(circuits)} circuit(s) from {path.name}")
    print(f"Site columns: {len(layout.sites)}")
    print(f"Written to {out_path}")
    print()
    print(text, end="")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
