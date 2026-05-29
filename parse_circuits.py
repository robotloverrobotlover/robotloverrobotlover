import openpyxl
import sys
import os

def get_cell_color(cell):
    try:
        fill = cell.fill
        if not fill or fill.fill_type in (None, "none"):
            return None
        fg = fill.fgColor
        rgb = fg.rgb
        if not isinstance(rgb, str):
            try:
                rgb = "{:02X}{:02X}{:02X}{:02X}".format(rgb.alpha, rgb.red, rgb.green, rgb.blue)
            except Exception:
                rgb = str(rgb)
        rgb = rgb.upper()
        if rgb in ("00000000", "FF000000", "FFFFFFFF", "00FFFFFF"):
            return None
        color_map = {
            "FF2E7D32": "green",     "FF00695C": "teal",     "FF1565C0": "blue",
            "FF00AA00": "green",     "FF007070": "teal",     "FF005500": "dark green",
            "FF008000": "green",     "FF006400": "dark green",
            "FF4E7C2F": "mid green", "FF1F5C1F": "dark green",
            "FF6B8E23": "olive",     "FF90C060": "light green",
        }
        return color_map.get(rgb, f"#{rgb[-6:]}")
    except Exception:
        return None


def load_sheet(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = None
    for name in wb.sheetnames:
        if "mpls" in name.lower() or "pre" in name.lower():
            sheet = wb[name]; break
    return sheet or wb.active


def diagnose(filepath):
    sheet = load_sheet(filepath)
    print(f"Sheet: '{sheet.title}'  (max_row={sheet.max_row}, max_col={sheet.max_column})")
    print("\n--- First 6 rows (all non-empty cells) ---")
    for i, row in enumerate(sheet.iter_rows(max_row=6), 1):
        cells = [(c.column, repr(c.value)) for c in row if c.value is not None]
        if cells:
            print(f"  Row {i}: {cells}")
    print("\n--- Merged ranges (first 15) ---")
    for j, m in enumerate(sheet.merged_cells.ranges):
        if j >= 15: break
        print(f"  {m}")
    print("\n--- Sample data rows 3-8 (all non-empty cells) ---")
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=8), 3):
        cells = [(c.column, repr(c.value)) for c in row if c.value is not None]
        if cells:
            print(f"  Row {i}: {cells}")


def parse_excel(filepath):
    sheet = load_sheet(filepath)

    # Build merged-cell map
    merged_map = {}
    for rng in sheet.merged_cells.ranges:
        tl = (rng.min_row, rng.min_col)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merged_map[(r, c)] = tl

    def real_value(row_1, col_1):
        key = (row_1, col_1)
        if key in merged_map:
            tr, tc = merged_map[key]
            return sheet.cell(tr, tc).value
        return sheet.cell(row_1, col_1).value

    # Find header row
    header_row = None
    for row in sheet.iter_rows():
        for cell in row:
            if "full/partial" in str(cell.value or "").lower():
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        raise ValueError("Could not find header row (looking for 'Full/Partial?'). Run --diagnose.")

    # Parse fixed columns and site columns from header row
    col_fp = col_comm = col_circ = col_new = col_name = col_email = None
    col_ac = col_mpls = None
    site_headers = []   # (col_1indexed, site_name)
    last_fixed_col = 0

    for cell in sheet[header_row]:
        v = str(cell.value or "").strip()
        vl = v.lower()
        c = cell.column
        if "full/partial" in vl:                           col_fp    = c
        elif "comments"   in vl:                           col_comm  = c
        elif "circuit #"  in vl and "new" not in vl:      col_circ  = c
        elif "new circuit" in vl:                          col_new   = c
        elif "circuit name" in vl:                         col_name  = c
        elif "email" in vl or "noc" in vl or "cau" in vl: col_email = c
        elif "ac:" in vl or vl == "ac":                    col_ac    = c
        elif "mpls" in vl or "mux" in vl:                 col_mpls  = c

    # Site columns = everything between col_name and col_ac (or col_mpls)
    # i.e. all header columns after the last known fixed col, before AC/MPLS
    known_fixed = {col_fp, col_comm, col_circ, col_new, col_name, col_email,
                   col_ac, col_mpls} - {None}
    last_fixed_col = max((c for c in [col_fp, col_comm, col_circ, col_new,
                                       col_name, col_email] if c), default=1)
    stop_col = min((c for c in [col_ac, col_mpls] if c),
                   default=sheet.max_column + 1)

    for cell in sheet[header_row]:
        c = cell.column
        v = str(cell.value or "").strip()
        if v and c > last_fixed_col and c < stop_col and c not in known_fixed:
            site_headers.append((c, v))

    print(f"Header row    : {header_row}")
    print(f"Fixed columns : Full/Partial={col_fp}, Circuit#={col_circ}, "
          f"NewCircuit#={col_new}, Name={col_name}, Email={col_email}")
    print(f"Site columns  : {len(site_headers)} found — "
          f"{[name for _, name in site_headers[:6]]}{'...' if len(site_headers) > 6 else ''}")

    if not site_headers:
        raise ValueError("No site columns found between Circuit Name and AC columns. Run --diagnose.")

    # Skip rows merged into header
    first_data = header_row + 1
    if col_circ:
        for r in range(header_row + 1, sheet.max_row + 1):
            key = (r, col_circ)
            if key in merged_map and merged_map[key][0] <= header_row:
                first_data = r + 1
            else:
                first_data = r
                break

    print(f"First data row: {first_data}\n")

    circuits = []
    visited = set()
    r = first_data

    while r <= sheet.max_row:
        if r in visited:
            r += 1; continue

        circ_val = str(real_value(r, col_circ) or "").strip() if col_circ else ""
        name_val = str(real_value(r, col_name) or "").strip() if col_name else ""

        if not circ_val and not name_val:
            r += 1; continue

        # Skip if merged up into header zone
        if col_circ:
            key = (r, col_circ)
            if key in merged_map and merged_map[key][0] < first_data:
                r += 1; continue

        # Find how many rows this circuit spans (merged fixed cols)
        span_end = r
        if col_circ:
            for rr in range(r + 1, sheet.max_row + 1):
                k = (rr, col_circ)
                if k in merged_map and merged_map[k][0] == r:
                    span_end = rr
                else:
                    break

        def cv(col):
            if col is None: return ""
            return str(real_value(r, col) or "").strip()

        fp    = cv(col_fp)
        comm  = cv(col_comm)
        circ  = cv(col_circ)
        new_c = cv(col_new)
        name  = cv(col_name)
        email = cv(col_email)
        ac    = cv(col_ac)   if col_ac   else ""
        mpls  = cv(col_mpls) if col_mpls else ""

        # Collect connections row by row
        connections = []
        for rr in range(r, span_end + 1):
            row_sites = []
            for col_1, site_name in site_headers:
                cell = sheet.cell(rr, col_1)
                v = str(cell.value or "").strip()
                if v:
                    color = get_cell_color(cell)
                    row_sites.append((site_name, v, color))
            if row_sites:
                connections.append(row_sites)

        visited.update(range(r, span_end + 1))

        circuits.append({
            "full_partial": fp, "comments": comm,
            "circuit": circ,    "new_circuit": new_c,
            "name": name,       "email": email,
            "ac": ac,           "mpls_reuse": mpls,
            "connections": connections,
        })

        r = span_end + 1

    return circuits


def format_output(circuits):
    lines = []
    for i, c in enumerate(circuits, 1):
        lines.append(f"Block {i}")
        lines.append(f"1. Circuit #:     {c['circuit'] or '(blank)'}")
        lines.append(f"2. New Circuit #: {c['new_circuit'] or '(blank)'}")
        lines.append(f"3. Circuit Name:  {c['name'] or '(blank)'}")

        connections = c["connections"]
        if not connections:
            lines.append("   (No site connections found)")
        else:
            for conn_idx, row_sites in enumerate(connections, 1):
                lines.append(f"   Connection {conn_idx}:")
                for site_name, port_val, color in row_sites:
                    color_str = f", color: {color}" if color else ""
                    lines.append(f"      {site_name} (port: {port_val}{color_str})")
                if conn_idx < len(connections):
                    lines.append("")

        extras = []
        if c["full_partial"]: extras.append(f"Full/Partial: {c['full_partial']}")
        if c["comments"]:     extras.append(f"Comments: {c['comments']}")
        if c["email"]:        extras.append(f"NOC/Email: {c['email']}")
        if c["ac"]:           extras.append(f"AC: {c['ac']}")
        if c["mpls_reuse"]:   extras.append(f"MPLS Port Reuse: {c['mpls_reuse']}")
        if extras:
            lines.append("   [" + " | ".join(extras) + "]")
        lines.append("")
    return "\n".join(lines)


# ── Main ─────────────────────────────────────────────────────────────────────
if len(sys.argv) < 2:
    print("Usage:")
    print("  python parse_circuits.py <input.xlsx> [output.txt]")
    print("  python parse_circuits.py <input.xlsx> --diagnose")
    sys.exit(1)

input_file = sys.argv[1]
if not os.path.exists(input_file):
    print(f"Error: File not found: {input_file}")
    sys.exit(1)

if "--diagnose" in sys.argv:
    diagnose(input_file)
    sys.exit(0)

output_file = sys.argv[2] if len(sys.argv) >= 3 else os.path.splitext(input_file)[0] + "_output.txt"

print(f"Reading: {input_file}")
circuits = parse_excel(input_file)
print(f"Found {len(circuits)} circuit(s).\n")

text = format_output(circuits)
print(text)

with open(output_file, "w", encoding="utf-8") as f:
    f.write(text)

print(f"Output saved to: {output_file}")
